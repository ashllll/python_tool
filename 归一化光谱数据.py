import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.spatial.distance import euclidean
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class SpectrumAnalyzer(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        self.label = QLabel('请选择一个光谱数据文件', self)
        layout.addWidget(self.label)

        self.btn = QPushButton('选择文件', self)
        self.btn.clicked.connect(self.showFileDialog)
        layout.addWidget(self.btn)

        self.result_label = QLabel('', self)
        layout.addWidget(self.result_label)

        self.setLayout(layout)
        self.setWindowTitle('光谱分析器')
        self.show()

    def showFileDialog(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, '选择光谱数据文件', '', 'Excel Files (*.xlsx *.xls)', options=options)
        if file_path:
            analysis_result = self.analyze_spectrum(file_path)
            if analysis_result is not None:
                results, diff_info = analysis_result
                self.save_results_to_excel(file_path, results, diff_info)
                self.result_label.setText('光谱分析计算完成并保存到文件。')
            else:
                self.result_label.setText('无法进行光谱分析，请检查文件格式和内容。')
        else:
            self.result_label.setText('未选择文件。')

    def analyze_spectrum(self, file_path):
        try:
            # 读取Excel中的光谱数据，第一行作为列名
            spectrum_data = pd.read_excel(file_path, sheet_name='Sheet1', header=0)

            # 获取波长信息
            wavelengths = spectrum_data.iloc[:, 0]
            results = {}
            normalized_spectrum_data = spectrum_data.copy()

            # 存储归一化后的光谱数据
            normalized_columns = {}

            # 遍历每一列的光谱数据，从第二列开始
            for column in spectrum_data.columns[1:]:
                intensities = spectrum_data[column]

                # 清洗数据，移除 NaN 和 inf 值
                intensities = self.clean_data(intensities)

                # 归一化光谱数据
                normalized_intensities = self.normalize_spectrum(intensities)
                normalized_intensities = self.clean_data(normalized_intensities)  # 再次清洗归一化后的数据
                normalized_spectrum_data[column] = normalized_intensities
                normalized_columns[column] = normalized_intensities

                # 计算蓝光占比
                blue_light_ratio = self.calculate_blue_light_ratio(wavelengths, normalized_intensities)
                results[f'{column} 蓝光占比'] = blue_light_ratio

            # 计算归一化光谱之间的差异
            diff_info = self.calculate_differences(normalized_columns)

            # 将归一化光谱数据添加到结果中
            results['normalized_spectrum_data'] = normalized_spectrum_data

            return results, diff_info

        except Exception as e:
            print(f"错误: {e}")
            return None

    def clean_data(self, data):
        """清洗数据，移除 NaN 和 inf 值"""
        data = np.where(np.isfinite(data), data, 0)
        return data

    def normalize_spectrum(self, intensities):
        min_intensity = np.min(intensities)
        max_intensity = np.max(intensities)
        normalized_intensities = (intensities - min_intensity) / (max_intensity - min_intensity)
        return normalized_intensities

    def calculate_blue_light_ratio(self, wavelengths, intensities):
        blue_light_data = (wavelengths >= 400) & (wavelengths <= 500)
        blue_light_intensity = np.trapz(intensities[blue_light_data], wavelengths[blue_light_data])
        total_intensity = np.trapz(intensities, wavelengths)
        blue_light_ratio = blue_light_intensity / total_intensity
        return blue_light_ratio

    def calculate_differences(self, normalized_columns):
        columns = list(normalized_columns.keys())
        diff_info = []

        for i in range(len(columns)):
            for j in range(i + 1, len(columns)):
                col1 = columns[i]
                col2 = columns[j]
                diff = euclidean(normalized_columns[col1], normalized_columns[col2])
                diff_info.append((col1, col2, diff))

        # 标注差异较大的光谱数据
        diff_info.sort(key=lambda x: x[2], reverse=True)  # 按差异值降序排列
        return diff_info

    def save_results_to_excel(self, file_path, results, diff_info):
        normalized_spectrum_data = results.pop('normalized_spectrum_data')

        results_df = pd.DataFrame(list(results.items()), columns=['Property', 'Value'])
        diff_df = pd.DataFrame(diff_info, columns=['Spectrum 1', 'Spectrum 2', 'Difference'])

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            results_df.to_excel(writer, sheet_name='Analysis Results', index=False)
            normalized_spectrum_data.to_excel(writer, sheet_name='Normalized Spectrum Data', index=False)
            diff_df.to_excel(writer, sheet_name='Differences', index=False)

        self.plot_differences(diff_info, file_path)

    def plot_differences(self, diff_info, file_path):
        diff_df = pd.DataFrame(diff_info, columns=['Spectrum 1', 'Spectrum 2', 'Difference'])
        plt.figure(figsize=(10, 6))
        plt.bar(diff_df.index, diff_df['Difference'])
        plt.xlabel('Comparison Pairs')
        plt.ylabel('Difference')
        plt.title('Differences between Normalized Spectra')
        plt.tight_layout()

        plot_image_path = 'differences_plot.png'
        plt.savefig(plot_image_path)
        plt.close()

        wb = load_workbook(file_path)
        ws = wb.create_sheet('Differences Plot')
        img = Image(plot_image_path)
        ws.add_image(img, 'A1')
        wb.save(file_path)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = SpectrumAnalyzer()
    sys.exit(app.exec_())