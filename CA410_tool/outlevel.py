import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import PatternFill


def get_file_path():
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    file_path = filedialog.askopenfilename(title="选择要处理的CSV文件", filetypes=[("CSV files", "*.csv")])
    if not file_path:
        raise Exception("必须选择一个CSV文件")
    return file_path


def calculate_and_save_results(file_path):
    # 读取CSV文件
    df = pd.read_csv(file_path)

    # 检查是否存在需要的列
    if 'Lv' not in df.columns:
        raise Exception("结果文件中没有找到 'Lv' 列")

    # 执行计算，并将结果保存到新列
    df['Detal L/L'] = df['Lv'].diff() / df['Lv'].shift(1)

    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Measurement Results"

    # 将数据帧转换为行，并添加到工作表中
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # 遍历E列数据，大于0.005的单元格标记为红色
    red_fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, max_row=ws.max_row):
        for cell in row:
            if cell.value > 0.005:
                cell.fill = red_fill

    # 创建一个折线图
    chart = LineChart()
    chart.title = "Lv"
    chart.style = 10
    chart.y_axis.title = 'Lv'
    chart.x_axis.title = 'Backlight'

    # 添加数据到图表
    data = Reference(ws, min_col=4, min_row=2, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    series = Series(data, title_from_data=True)
    chart.series.append(series)

    # 设置线条样式为细线
    series.graphicalProperties.line.width = 10000  # 单位是EMU，1点 = 12700 EMUs

    chart.set_categories(categories)
    ws.add_chart(chart, "F2")

    # 保存结果到Excel文件
    output_excel_path = os.path.splitext(file_path)[0] + '_updated.xlsx'
    wb.save(output_excel_path)
    print(f"计算结果已保存到: {output_excel_path}")


def main():
    # 获取CSV文件路径
    file_path = get_file_path()
    print(f"选择的文件是: {file_path}")

    # 执行计算并保存结果
    calculate_and_save_results(file_path)


if __name__ == "__main__":
    main()
