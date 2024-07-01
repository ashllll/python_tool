import sys
import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from PyQt5.QtWidgets import QApplication, QFileDialog

def normalize(data):
    lv = data['Lv']
    max_val = lv.max()
    min_val = lv.min()
    norm_lv = (lv - min_val) / (max_val - min_val)
    return norm_lv

def calc_transition(data):
    transition = data['Norm_Lv'].diff().abs()
    transition.iloc[0] = 0  # 第一个元素的变化为0
    return transition

def normalize_transition(data):
    trans = data['Transition']
    max_val = trans.max()
    min_val = trans.min()
    norm_trans = (trans - min_val) / (max_val - min_val)
    return norm_trans

def calc_uniformity(data):
    mean_lv = data['Lv'].mean()
    std_lv = data['Lv'].std()
    uniformity = std_lv / mean_lv
    return uniformity

def calc_delta_e(data1, data2):
    delta_e = np.sqrt((data1['x'] - data2['x'])**2 + (data1['y'] - data2['y'])**2)
    return delta_e

def calc_contrast(data):
    max_lv = data['Lv'].max()
    min_lv = data['Lv'].min()
    contrast = max_lv / min_lv
    return contrast

def calc_color_temperature(data):
    x = data['x'].mean()
    y = data['y'].mean()
    n = (x - 0.3320) / (0.1858 - y)
    cct = 449 * (n ** 3) + 3525 * (n ** 2) + 6823.3 * n + 5520.33
    return cct

def main():
    app = QApplication(sys.argv)
    options = QFileDialog.Options()
    options |= QFileDialog.ReadOnly
    file_path, _ = QFileDialog.getOpenFileName(None, "Select Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
    if not file_path:
        return

    df = pd.read_excel(file_path, header=2)

    W_data = df.iloc[:, 0:3]
    R_data = df.iloc[:, 3:6]
    G_data = df.iloc[:, 6:9]
    B_data = df.iloc[:, 9:12]

    W_data.columns = ['Lv', 'x', 'y']
    R_data.columns = ['Lv', 'x', 'y']
    G_data.columns = ['Lv', 'x', 'y']
    B_data.columns = ['Lv', 'x', 'y']

    W_data['Norm_Lv'] = normalize(W_data)
    R_data['Norm_Lv'] = normalize(R_data)
    G_data['Norm_Lv'] = normalize(G_data)
    B_data['Norm_Lv'] = normalize(B_data)

    W_data['Transition'] = calc_transition(W_data)
    R_data['Transition'] = calc_transition(R_data)
    G_data['Transition'] = calc_transition(G_data)
    B_data['Transition'] = calc_transition(B_data)

    W_data['Norm_Transition'] = normalize_transition(W_data)
    R_data['Norm_Transition'] = normalize_transition(R_data)
    G_data['Norm_Transition'] = normalize_transition(G_data)
    B_data['Norm_Transition'] = normalize_transition(B_data)

    W_uniformity = calc_uniformity(W_data)
    R_uniformity = calc_uniformity(R_data)
    G_uniformity = calc_uniformity(G_data)
    B_uniformity = calc_uniformity(B_data)

    delta_e_wr = calc_delta_e(W_data, R_data)
    delta_e_wg = calc_delta_e(W_data, G_data)
    delta_e_wb = calc_delta_e(W_data, B_data)

    W_contrast = calc_contrast(W_data)
    R_contrast = calc_contrast(R_data)
    G_contrast = calc_contrast(G_data)
    B_contrast = calc_contrast(B_data)

    W_cct = calc_color_temperature(W_data)
    R_cct = calc_color_temperature(R_data)
    G_cct = calc_color_temperature(G_data)
    B_cct = calc_color_temperature(B_data)

    output_file = os.path.join(os.path.dirname(file_path), 'output_data.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        W_data.to_excel(writer, sheet_name='W_Data', index=False)
        R_data.to_excel(writer, sheet_name='R_Data', index=False)
        G_data.to_excel(writer, sheet_name='G_Data', index=False)
        B_data.to_excel(writer, sheet_name='B_Data', index=False)

        summary = pd.DataFrame({
            'Uniformity': [W_uniformity, R_uniformity, G_uniformity, B_uniformity],
            'Contrast': [W_contrast, R_contrast, G_contrast, B_contrast],
            'CCT': [W_cct, R_cct, G_cct, B_cct]
        }, index=['W', 'R', 'G', 'B'])

        summary.to_excel(writer, sheet_name='Summary')

        delta_e_df = pd.DataFrame({
            'ΔE_W-R': delta_e_wr,
            'ΔE_W-G': delta_e_wg,
            'ΔE_W-B': delta_e_wb
        })

        delta_e_df.to_excel(writer, sheet_name='Delta_E', index=False)

    wb = load_workbook(output_file)
    ws = wb['Delta_E']

    red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    threshold = 3
    for col in ['A', 'B', 'C']:
        ws.conditional_formatting.add(f'{col}2:{col}{len(delta_e_df) + 1}', 
                                      CellIsRule(operator='greaterThan', formula=[str(threshold)], fill=red_fill))

    wb.save(output_file)

    plt.figure(figsize=(12, 8))
    plt.plot(W_data['Norm_Lv'], label='W Norm Lv')
    plt.plot(R_data['Norm_Lv'], label='R Norm Lv')
    plt.plot(G_data['Norm_Lv'], label='G Norm Lv')
    plt.plot(B_data['Norm_Lv'], label='B Norm Lv')

    plt.xlabel('Index')
    plt.ylabel('Normalized Lv')
    plt.title('Normalized Lv of WRGB')
    plt.legend()
    plt.grid(True)
    plt.savefig(os.path.join(os.path.dirname(file_path), 'normalized_lv_plot.png'))
    plt.show()

    plt.figure(figsize=(12, 8))
    plt.plot(W_data['Norm_Transition'], label='W Norm Transition', linestyle='--')
    plt.plot(R_data['Norm_Transition'], label='R Norm Transition', linestyle='--')
    plt.plot(G_data['Norm_Transition'], label='G Norm Transition', linestyle='--')
    plt.plot(B_data['Norm_Transition'], label='B Norm Transition', linestyle='--')

    plt.xlabel('Index')
    plt.ylabel('Normalized Transition')
    plt.title('Normalized Transition of WRGB')
    plt.legend()
    plt.grid(True)
    plt.savefig(os.path.join(os.path.dirname(file_path), 'normalized_transition_plot.png'))
    plt.show()

    plt.figure(figsize=(12, 8))
    plt.plot(delta_e_wr, label='ΔE W-R')
    plt.plot(delta_e_wg, label='ΔE W-G')
    plt.plot(delta_e_wb, label='ΔE W-B')

    plt.xlabel('Index')
    plt.ylabel('ΔE')
    plt.title('Color Difference (ΔE) of WRGB')
    plt.legend()
    plt.grid(True)
    plt.savefig(os.path.join(os.path.dirname(file_path), 'delta_e_plot.png'))
    plt.show()

    plt.figure(figsize=(12, 8))
    plt.plot(W_data['Lv'], label='W Lv')

    gamma_values = [1.9, 2.2, 2.4]
    colors = ['green', 'blue', 'red']
    labels = ['Lower Limit (1.9)', 'Standard (2.2)', 'Upper Limit (2.4)']
    for gamma, color, label in zip(gamma_values, colors, labels):
        reference_curve = np.linspace(0, 1, len(W_data)) ** gamma
        plt.plot(reference_curve * W_data['Lv'].max(), color=color, linestyle='--', label=label)

    plt.xlabel('Index')
    plt.ylabel('Lv')
    plt.title('W Gamma Curve')
    plt.legend()
    plt.grid(True)
    plt.savefig(os.path.join(os.path.dirname(file_path), 'w_gamma_curve.png'))
    plt.show()

if __name__ == "__main__":
    main()